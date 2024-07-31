from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import IntegrityError
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from datetime import datetime
from django.core.paginator import Paginator
from io import BytesIO
from decimal import Decimal
from django.core.files.base import ContentFile
from django.templatetags.static import static
import os
import calendar
from django.conf import settings
from copy import deepcopy, copy
from django.contrib import messages
from django.http import HttpResponseRedirect
import requests
from datetime import timedelta
from django.utils import timezone
import csv
from itertools import islice
from django.http import JsonResponse
from django.db.models import OuterRef, Subquery
from io import TextIOWrapper
from django.db import transaction
from django.http import HttpResponseNotFound
from django.urls import reverse_lazy
import time
from collections import defaultdict
from django.db.models import F
from weasyprint import HTML
from django.template.loader import render_to_string
# Importe de formularios

# Importe de modelos
from .models import CoberturaInnominada, CoberturaNominada, AseguradoCredito, Vencimiento, Localidad, Flota, VehiculoFlota, VehiculoInfoAuto, MarcaInfoAuto, PrecioAnual, Movimiento, TarifaFlota, Cliente, AccessToken, RefreshToken, Localidad

# Importe de librerias
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

import xlwings as xw
from unidecode import unidecode

# Importe de funciones
from .api_auth import ApiAuthentication, AuthenticationError
from .api_manager import ApiManager
from .utils import get_tarifas, get_vehicle_type, convert_tipo_cobertura, convert_date, handle_aumento_suma_asegurada, handle_baja_items, handle_cambio_cobertura, handle_modificacion_datos, handle_renovacion_alta_items
from .utils import importar_datos_roemmers_saicf, importar_datos_roemmers_alberto_guillermo, importar_datos_rofina_saicf, importar_datos_ganadera_santa_isabel, comparar_totales
from .utils_creditos import cargar_datos_innominados, cargar_datos_nominados, obtener_datos_solicitudes_cobertura, obtener_datos_clientes_sin_cobertura

# Roles y permisos
def is_staff_user(user):
    return user.is_staff

def permiso_basico(user):
    return user.groups.filter(name='PermisoBasico').exists()

def permiso_avanzado(user):
    return user.groups.filter(name='PermisoAvanzado').exists()

staff_required = user_passes_test(is_staff_user, login_url=reverse_lazy('signin'))

# 404 page
def pagina_no_encontrada(request, exception):
    print("Error 404 ocurrido")
    return HttpResponseNotFound(render(request, '404.html'))

class HomeView(View):
    def get(self, request, *args, **kwargs):
        context = {
            
        }
        return redirect('login')

# Inicio
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class InicioView(View):
    def get(self, request, *args, **kwargs):
        
        context = {
            
        }
        return render(request, 'dashboard.html', context)
    def post(self, request, *args, **kwargs):
        lista_errores = []
        context = {
            'errores': lista_errores
        }
        if 'descargar_excel' in request.POST:
            # Nombre del archivo que quieres descargar
            file_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel', 'modelo_ejemplo.xlsx')


            # Abre el archivo y lee su contenido
            with open(file_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=modelo_ejemplo.xlsx'
                return response
        if 'importar_localidades' in request.POST:
            file1 = request.FILES.get('file1')

            # Envuelve el archivo en un TextIOWrapper para manejar la codificación
            csv_file_wrapper = TextIOWrapper(file1.file, encoding='utf-8')

            # Usa DictReader para obtener un diccionario por fila
            csv_reader = csv.DictReader(csv_file_wrapper)

            # Itera sobre las filas del CSV
            for row_number, row in enumerate(csv_reader, start=2):
                municipio = row.get('municipio_nombre', '')
                localidad = row.get('nombre', '')
                provincia = row.get('provincia_nombre', '')
                zona = row.get('zona_nombre', '')

                # Crea el objeto Localidad
                Localidad.objects.create(
                    nombre_localidad=localidad,
                    nombre_municipio=municipio,
                    nombre_provincia=provincia,
                    zona=zona
                )

        if "calcular_excel" in request.POST:
            file1 = request.FILES.get('file1')
            workbook = openpyxl.load_workbook(file1)
            sheet = workbook.active
            for row_number, (marca, modelo, tipo_vehiculo, patente, anio, okm, importado, zona, fecha_operacion, fecha_vigencia, operacion, cobertura, suma_asegurada, _, _, _, _) in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_values = sheet.cell(row=row_number, column=1).value
                if row_values is None:
                    # Salir del bucle si la fila está vacía
                    break
                anio_vehiculo = anio
                anio_actual = datetime.now().year
                antiguedad_vehiculo = anio_actual - anio_vehiculo
                
                # Mapeo de antigüedad a categoría
                if antiguedad_vehiculo > 10:
                    antiguedad_categoria = "MÁS DE 10"
                elif 6 <= antiguedad_vehiculo <= 10:
                    antiguedad_categoria = "6 A 10"
                else:
                    antiguedad_categoria = "5"
                
                try:    
                    tarifa = TarifaFlota.objects.get(
                        tipo_vehiculo=tipo_vehiculo,
                        antiguedad=antiguedad_categoria,
                        zona__icontains=zona,
                        tipo_cobertura__contains=cobertura,
                    )
                except: 
                    error_message = f"No se encontró tarifa para {tipo_vehiculo}, {antiguedad_categoria}, {zona}, {cobertura}"
                    print(error_message)
                    lista_errores.append(error_message)  # Agregar el mensaje a una lista de errores
                tasa = tarifa.tasa
                prima_rc_anual = tarifa.prima_rc_anual

                prima_anual = (suma_asegurada * (tasa / 1000)) + prima_rc_anual

                # Calcular los días de vigencia
                dias_vigencia = (fecha_vigencia - fecha_operacion).days

                prima_vigente = prima_anual * dias_vigencia / 365
                

                derecho_emision = 2400
                recargo_financiero = Decimal('5.68')
                cobertura_nacional = 75000
                cobertura_importado = 112500

                # Determinar la cobertura según si la unidad es importada o no
                cobertura = cobertura_importado if importado == "SI" else cobertura_nacional

                premio_anual = prima_anual + cobertura + derecho_emision + ((prima_anual * recargo_financiero) / 100)
                premio_vigente = prima_vigente + cobertura + derecho_emision + ((prima_vigente * recargo_financiero) / 100)

                if operacion == "BAJA":
                    prima_anual = -prima_anual
                    prima_vigente = -prima_vigente
                    premio_anual = -premio_anual
                    premio_vigente = -premio_vigente

                # Redondear valores
                premio_anual_redondeado = round(premio_anual, 2)
                premio_vigente_redondeado = round(premio_vigente, 2)
                prima_vigente_redondeada = round(prima_vigente, 2)
                # Actualizar los valores en las columnas existentes
                sheet.cell(row=row_number, column=sheet.max_column - 3, value=prima_anual)  # Actualizar la columna de Prima Anual
                sheet.cell(row=row_number, column=sheet.max_column - 2, value=prima_vigente_redondeada)  # Actualizar la columna de Prima Vigente
                sheet.cell(row=row_number, column=sheet.max_column - 1, value=premio_anual_redondeado)  # Actualizar la columna de Prremio Anual
                sheet.cell(row=row_number, column=sheet.max_column, value=premio_vigente_redondeado)  # Actualizar la columna de Premio Vigente
                        
                
                # Guardar la hoja de cálculo actualizada
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Crear una respuesta HTTP con el archivo adjunto
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=resultados_actualizados.xlsx'

            return response
        return render(request, 'dashboard.html', context)

# Clientes
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class ClientesView(View):
    def get(self, request, *args, **kwargs):
        clientes = Cliente.objects.all()
        
        clientes_paginados = Paginator(clientes, 30)
        page_number = request.GET.get("page")
        filter_pages = clientes_paginados.get_page(page_number)

        context = {
            'clientes': clientes, 
            'pages': filter_pages,

        }
        return render(request, 'clientes/clientes.html', context)
    def post(self, request, *args, **kwargs):
        # Obtén los datos del formulario directamente desde request.POST
        
        nombre = request.POST.get('nombre')
        cuit = request.POST.get('cuit')
        nacionalidad = request.POST.get('nacionalidad')
        provincia = request.POST.get('provincia')
        localidad = request.POST.get('localidad')
        direccion = request.POST.get('direccion')
        telefono = request.POST.get('telefono')
        email = request.POST.get('email')
        recargo_financiero = request.POST.get('recargo_financiero')
        impuestos_y_sellados = request.POST.get('imp_y_sellados')
        iibb = request.POST.get('iibb')
        iva = request.POST.get('iva')
        

        # Crea una nueva instancia de Flota
        nuevo_cliente = Cliente(
            
            nombre_cliente=nombre,
            cuit=cuit,
            nacionalidad=nacionalidad,
            provincia=provincia,
            localidad=localidad,
            direccion=direccion,
            telefono=telefono,
            email=email,
            recargo_financiero=recargo_financiero,
            imp_y_sellados=impuestos_y_sellados,
            iibb=iibb,
            iva=iva,
            
        )
        try:
            # Intenta crear el nuevo elemento
            nuevo_cliente.save()
            messages.success(request, 'El elemento se creó exitosamente.')
        except Exception as e:
            # Si hay un error al crear el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')

        # Redirige, incluyendo los mensajes en el contexto
        return HttpResponseRedirect(request.path_info)
    
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class DetalleClienteView(View):
    def get(self, request, cliente_id):
        cliente = get_object_or_404(Cliente, id=cliente_id)
        # Formatear los decimales para poder mostrarlos
        if cliente.recargo_financiero:
            recargo_financiero_formatted = "{:.3f}".format(cliente.recargo_financiero).replace(',', '.')
        else:
            recargo_financiero_formatted = 0
        if cliente.imp_y_sellados:
            imp_y_sellados_formatted = "{:.3f}".format(cliente.imp_y_sellados).replace(',', '.')
        else:
            imp_y_sellados_formatted = 0
        if cliente.iibb:
            iibb_formatted = "{:.3f}".format(cliente.iibb).replace(',', '.')
        else:
            iibb_formatted = 0
        print(imp_y_sellados_formatted)
        iva_formatted = "{:.3f}".format(cliente.iva).replace(',', '.')
    
        context = {
            'cliente': cliente,
            'recargo_financiero_formatted': recargo_financiero_formatted,
            'imp_y_sellados_formatted': imp_y_sellados_formatted,
            'iibb_formatted': iibb_formatted,
            'iva_formatted': iva_formatted,
        }
        return render(request, 'clientes/detalle_cliente.html', context)

    def post(self, request, cliente_id):
        cliente = get_object_or_404(Cliente, id=cliente_id)
        # Obtén los datos del formulario directamente desde request.POST
        nombre = request.POST.get('nombre')
        cuit = request.POST.get('cuit')
        nacionalidad = request.POST.get('nacionalidad')
        provincia = request.POST.get('provincia')
        localidad = request.POST.get('localidad')
        direccion = request.POST.get('direccion')
        telefono = request.POST.get('telefono')
        email = request.POST.get('email')
        recargo_financiero = request.POST.get('rf')
        imp_y_sellados = request.POST.get('imp_y_sellados')
        iibb = request.POST.get('iibb')
        iva = request.POST.get('iva')
        
        # Actualiza los campos de la tarifa con los datos del formulario
        cliente.nombre_cliente = nombre
        cliente.cuit = cuit
        cliente.nacionalidad = nacionalidad
        cliente.provincia = provincia
        cliente.localidad = localidad
        cliente.direccion = direccion
        cliente.telefono = telefono
        cliente.email = email
        cliente.recargo_financiero = recargo_financiero
        cliente.imp_y_sellados = imp_y_sellados
        cliente.iibb = iibb
        cliente.iva = iva
        
        try:
            # Intenta guardar la actualización del elemento
            cliente.save()
            messages.success(request, 'El elemento se actualizó exitosamente.')
        except Exception as e:
            # Si hay un error al actualizar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo actualizar el elemento. Detalles: {str(e)}')
        return redirect('clientes')

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class EliminarClienteView(View):
    def get(self, request, cliente_id):
        cliente = get_object_or_404(Cliente, id=cliente_id)
        
        context = {
            'cliente': cliente,
        }
        return redirect('clientes')

    def post(self, request, cliente_id):
        cliente = get_object_or_404(Cliente, id=cliente_id)

        try:
            # Intenta guardar la eliminacion del elemento
            cliente.delete()
            messages.success(request, 'El elemento se eliminó exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar el elemento. Detalles: {str(e)}')
        
        
        return redirect('clientes')
# Movimientos
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class ObtenerDatosMovimientoView(View):
    def get(self, request, movimiento_id):
        movimiento = Movimiento.objects.get(id=movimiento_id)

        datos_movimiento = {
            'movimiento_id': movimiento_id,
            'numero_endoso': movimiento.numero_endoso,
            'motivo_endoso': movimiento.motivo_endoso,
            'numero_orden': movimiento.numero_orden,
            'fecha_alta_op': movimiento.fecha_alta_op,
            'porcentaje_dif_prima': "{:.3f}".format(movimiento.prima_pza_porcentaje_diferencia).replace(',', '.'),
            'porcentaje_dif_premio': "{:.3f}".format(movimiento.premio_con_iva_porcentaje_diferencia).replace(',', '.')
            # Agrega otros campos del movimiento según sea necesario
        }

        return JsonResponse(datos_movimiento)
    
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoAvanzado').exists() or user.is_staff), name='dispatch')
class EditarDatosMovimientoView(View):
    def post(self, request, flota_id, movimiento_id):
        movimiento = get_object_or_404(Movimiento, id=movimiento_id)
        numero_endoso = request.POST.get('numero_endoso')
        motivo_endoso = request.POST.get('motivo_endoso')
        numero_orden = request.POST.get('numero_orden')
        fecha_alta_op = request.POST.get('fecha_alta_op')

        movimiento.numero_endoso = numero_endoso
        movimiento.motivo_endoso = motivo_endoso
        movimiento.numero_orden = numero_orden
        movimiento.fecha_alta_op = fecha_alta_op

        try:
            # Intenta guardar
            movimiento.save()
            messages.success(request, 'El elemento se actualizó exitosamente.')
        except Exception as e:
            # Si hay un error captura la excepción
            messages.error(request, f'Error: No se pudo actualizar el elemento. Detalles: {str(e)}')

        return redirect('detalle_flota', flota_id=flota_id)

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class EliminarMovimientoView(View):
    def post(self, request, flota_id, movimiento_id):
        
        movimiento = get_object_or_404(Movimiento, id=movimiento_id)
        try:
            # Intenta guardar la eliminación del elemento
            movimiento.delete()
            messages.success(request, 'El elemento se eliminó exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar el elemento. Detalles: {str(e)}')
        
        return redirect('detalle_flota', flota_id=flota_id)

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class ExportarMovimientoView(View):
    def post(self, request, flota_id, movimiento_id):
        # Obtener movimiento
        flota = get_object_or_404(Flota, id=flota_id)
        movimiento = get_object_or_404(Movimiento, id=movimiento_id)
        # Nombre del archivo modelo
        file_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel', 'exportar_ult_estado.xlsx')
    
        # Obtener los vehículos del movimiento
        vehiculos = VehiculoFlota.history.filter(flota=flota, movimiento=movimiento)
        # Crear nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        print(vehiculos)
        
        # Estilos
        font = Font(name="Calibri", size=11, bold=False)
        font_bold = Font(name="Calibri", size=12, bold=True)
        font_header = Font(name="Calibri", size=12, bold=False, color="FFFFFF")
        relleno = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
        bordes = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))  # Bordes finos y negro
        relleno_header = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        
        # Obtener el número de la próxima fila disponible
        next_row = 10
        num_vehiculo = 1
        
        # Tamaños de filas y columnas
        sheet.row_dimensions[9].height = 30  # Establece la altura de la novena fila en 30 puntos
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        
        # Agregar "encabezado"
        sheet.cell(row=2, column=1, value=movimiento.flota.cliente.nombre_cliente).font = font_bold
        sheet.cell(row=4, column=1, value="POLIZA").font = font_bold
        sheet.cell(row=7, column=1, value='Endoso '+movimiento.numero_endoso).font = font_bold
        sheet.cell(row=9, column=1, value='RIESGO').font = font_header
        sheet.cell(row=9, column=2, value='MARCA').font = font_header
        sheet.cell(row=9, column=3, value='DESCRIPCION').font = font_header
        sheet.cell(row=9, column=4, value='AÑO').font = font_header
        sheet.cell(row=9, column=5, value='PATENTE').font = font_header
        sheet.cell(row=9, column=6, value='Premio sin IVA').font = font_header
        sheet.cell(row=9, column=7, value='Premio con IVA').font = font_header
        
        sheet.cell(row=4, column=2, value=movimiento.flota.poliza).font = font_bold
        sheet.cell(row=7, column=2, value=movimiento.motivo_endoso).font = font_bold
        sheet.cell(row=9, column=1).fill = relleno_header
        sheet.cell(row=9, column=2).fill = relleno_header
        sheet.cell(row=9, column=3).fill = relleno_header
        sheet.cell(row=9, column=4).fill = relleno_header
        sheet.cell(row=9, column=5).fill = relleno_header
        sheet.cell(row=9, column=6).fill = relleno_header
        sheet.cell(row=9, column=7).fill = relleno_header
        sheet.cell(row=9, column=1).border = bordes
        sheet.cell(row=9, column=2).border = bordes
        sheet.cell(row=9, column=3).border = bordes
        sheet.cell(row=9, column=4).border = bordes
        sheet.cell(row=9, column=5).border = bordes
        sheet.cell(row=9, column=6).border = bordes
        sheet.cell(row=9, column=7).border = bordes
        premio_sin_iva_total = 0
        premio_con_iva_total = 0
        # Iterar sobre los vehículos y agregar la información al archivo Excel existente
        for vehiculo in vehiculos:
            
            sheet.cell(row=next_row, column=1, value=num_vehiculo).font = font
            sheet.cell(row=next_row, column=2, value=vehiculo.marca).font = font
            sheet.cell(row=next_row, column=3, value=vehiculo.descripcion).font = font
            sheet.cell(row=next_row, column=4, value=vehiculo.anio).font = font
            sheet.cell(row=next_row, column=5, value=vehiculo.patente).font = font
            sheet.cell(row=next_row, column=6, value=vehiculo.premio_sin_iva).font = font
            sheet.cell(row=next_row, column=7, value=vehiculo.premio_con_iva).font = font
            sheet.cell(row=next_row, column=6).number_format = '#,##0'
            sheet.cell(row=next_row, column=7).number_format = '#,##0'
            # Aplicar color de relleno 
            sheet.cell(row=next_row, column=6).fill = relleno
            sheet.cell(row=next_row, column=7).fill = relleno
            
            # Aplicar bordes
            for col in range(1, 8):
                sheet.cell(row=next_row, column=col).border = bordes
            
            next_row += 1
            num_vehiculo += 1
            premio_sin_iva_total += vehiculo.premio_sin_iva
            premio_con_iva_total += vehiculo.premio_con_iva
            
        sheet.cell(row=next_row+1, column=6, value=premio_sin_iva_total).font = font
        sheet.cell(row=next_row+1, column=7, value=premio_con_iva_total).font = font
        sheet.cell(row=next_row+1, column=6).fill = relleno
        sheet.cell(row=next_row+1, column=7).fill = relleno
        sheet.cell(row=next_row+1, column=6).border = bordes
        sheet.cell(row=next_row+1, column=7).border = bordes
        sheet.cell(row=next_row+1, column=6).number_format = '#,##0'
        sheet.cell(row=next_row+1, column=7).number_format = '#,##0'
        # Crear una respuesta HTTP con el archivo Excel duplicado adjunto
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={movimiento.numero_endoso}.xlsx'
        workbook.save(response)
        workbook.close()

        return response
        
        return redirect('detalle_flota', flota_id=flota_id)

# Flotas

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class ExportarUltimoEstadoFlotaView(View):
    def post(self, request, flota_id):
        # Obtener flota y sus vehiculos
        flota = get_object_or_404(Flota, id=flota_id)
        vehiculos = VehiculoFlota.objects.filter(flota=flota)
        movimientos = Movimiento.objects.filter(flota=flota)
        # Nombre del archivo modelo
        file_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel', 'exportar_ult_estado.xlsx')
        
        # Crear nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'PZA 0'

        # Estilos
        font = Font(name="Calibri", size=11, bold=False)
        font_bold = Font(name="Calibri", size=12, bold=True)
        font_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        relleno = PatternFill(start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
        bordes = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))  # Bordes finos y negro
        relleno_header = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        # Obtener el número de la próxima fila disponible
        next_row = 10
        num_vehiculo = 1
        
        # Tamaños de filas y columnas
        sheet.row_dimensions[9].height = 30  # Establece la altura de la novena fila en 30 puntos
        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30

        
        # Agregar "encabezado"
        sheet.cell(row=2, column=1, value=flota.cliente.nombre_cliente).font = font_bold
        sheet.cell(row=4, column=1, value="POLIZA").font = font_bold

        sheet.cell(row=9, column=1, value='RIESGO').font = font_header
        sheet.cell(row=9, column=2, value='MARCA').font = font_header
        sheet.cell(row=9, column=3, value='DESCRIPCION').font = font_header
        sheet.cell(row=9, column=4, value='AÑO').font = font_header
        sheet.cell(row=9, column=5, value='PATENTE').font = font_header
        sheet.cell(row=9, column=6, value='Premio sin IVA').font = font_header
        sheet.cell(row=9, column=7, value='Premio con IVA').font = font_header
        
        
        sheet.cell(row=4, column=2, value=flota.poliza).font = font_bold
        
        sheet.cell(row=9, column=1).fill = relleno_header
        sheet.cell(row=9, column=2).fill = relleno_header
        sheet.cell(row=9, column=3).fill = relleno_header
        sheet.cell(row=9, column=4).fill = relleno_header
        sheet.cell(row=9, column=5).fill = relleno_header
        sheet.cell(row=9, column=6).fill = relleno_header
        sheet.cell(row=9, column=7).fill = relleno_header
        sheet.cell(row=9, column=1).border = bordes
        sheet.cell(row=9, column=2).border = bordes
        sheet.cell(row=9, column=3).border = bordes
        sheet.cell(row=9, column=4).border = bordes
        sheet.cell(row=9, column=5).border = bordes
        sheet.cell(row=9, column=6).border = bordes
        sheet.cell(row=9, column=7).border = bordes
        premio_sin_iva_total = 0
        premio_con_iva_total = 0
        # Iterar sobre los vehículos y agregar la información al archivo Excel existente
        for vehiculo in vehiculos:
            
            sheet.cell(row=next_row, column=1, value=vehiculo.id).font = font
            sheet.cell(row=next_row, column=2, value=vehiculo.marca).font = font
            sheet.cell(row=next_row, column=3, value=vehiculo.descripcion).font = font
            sheet.cell(row=next_row, column=4, value=vehiculo.anio).font = font
            sheet.cell(row=next_row, column=5, value=vehiculo.patente).font = font
            sheet.cell(row=next_row, column=6, value=vehiculo.premio_sin_iva).font = font
            sheet.cell(row=next_row, column=7, value=vehiculo.premio_con_iva).font = font
            sheet.cell(row=next_row, column=6).number_format = '#,##0'
            sheet.cell(row=next_row, column=7).number_format = '#,##0'
            
            # Aplicar color de relleno 
            sheet.cell(row=next_row, column=6).fill = relleno
            sheet.cell(row=next_row, column=7).fill = relleno
            
            # Aplicar bordes
            for col in range(1, 8):
                sheet.cell(row=next_row, column=col).border = bordes
            
            next_row += 1
            num_vehiculo += 1
            premio_sin_iva_total += vehiculo.premio_sin_iva
            premio_con_iva_total += vehiculo.premio_con_iva
            
        sheet.cell(row=next_row+1, column=6, value=premio_sin_iva_total).font = font
        sheet.cell(row=next_row+1, column=7, value=premio_con_iva_total).font = font
        sheet.cell(row=next_row+1, column=6).fill = relleno
        sheet.cell(row=next_row+1, column=7).fill = relleno
        sheet.cell(row=next_row+1, column=6).border = bordes
        sheet.cell(row=next_row+1, column=7).border = bordes
        sheet.cell(row=next_row+1, column=6).number_format = '#,##0'
        sheet.cell(row=next_row+1, column=7).number_format = '#,##0'
        # Iterar sobre los movimientos y crear una hoja para cada uno
        for movimiento in movimientos:
            hoja_movimiento = workbook.create_sheet(title=f'ENDOSO {movimiento.numero_endoso}')

            # Obtener vehículos asociados a este movimiento desde el historial
            vehiculos_movimiento = VehiculoFlota.history.filter(flota=flota, movimiento=movimiento)

            # Obtener el número de la próxima fila disponible
            next_row = 10
            num_vehiculo = 1
            
            # Tamaños de filas y columnas
            hoja_movimiento.row_dimensions[9].height = 30  # Establece la altura de la novena fila en 30 puntos
            hoja_movimiento.column_dimensions['A'].width = 14
            hoja_movimiento.column_dimensions['B'].width = 20
            hoja_movimiento.column_dimensions['C'].width = 30
            hoja_movimiento.column_dimensions['E'].width = 20
            hoja_movimiento.column_dimensions['F'].width = 30
            hoja_movimiento.column_dimensions['G'].width = 30

            
            # Agregar "encabezado"
            hoja_movimiento.cell(row=2, column=1, value=flota.cliente.nombre_cliente).font = font_bold
            hoja_movimiento.cell(row=4, column=1, value="POLIZA").font = font_bold
            hoja_movimiento.cell(row=7, column=1, value='Endoso '+movimiento.numero_endoso).font = font_bold
            hoja_movimiento.cell(row=9, column=1, value='RIESGO').font = font_header
            hoja_movimiento.cell(row=9, column=2, value='MARCA').font = font_header
            hoja_movimiento.cell(row=9, column=3, value='DESCRIPCION').font = font_header
            hoja_movimiento.cell(row=9, column=4, value='AÑO').font = font_header
            hoja_movimiento.cell(row=9, column=5, value='PATENTE').font = font_header
            hoja_movimiento.cell(row=9, column=6, value='Premio sin IVA').font = font_header
            hoja_movimiento.cell(row=9, column=7, value='Premio con IVA').font = font_header
            
            
            hoja_movimiento.cell(row=4, column=2, value=flota.poliza).font = font_bold
            hoja_movimiento.cell(row=7, column=2, value=movimiento.motivo_endoso).font = font_bold
            hoja_movimiento.cell(row=9, column=1).fill = relleno_header
            hoja_movimiento.cell(row=9, column=2).fill = relleno_header
            hoja_movimiento.cell(row=9, column=3).fill = relleno_header
            hoja_movimiento.cell(row=9, column=4).fill = relleno_header
            hoja_movimiento.cell(row=9, column=5).fill = relleno_header
            hoja_movimiento.cell(row=9, column=6).fill = relleno_header
            hoja_movimiento.cell(row=9, column=7).fill = relleno_header
            hoja_movimiento.cell(row=9, column=1).border = bordes
            hoja_movimiento.cell(row=9, column=2).border = bordes
            hoja_movimiento.cell(row=9, column=3).border = bordes
            hoja_movimiento.cell(row=9, column=4).border = bordes
            hoja_movimiento.cell(row=9, column=5).border = bordes
            hoja_movimiento.cell(row=9, column=6).border = bordes
            hoja_movimiento.cell(row=9, column=6).number_format = '#,##0'
            hoja_movimiento.cell(row=9, column=7).border = bordes
            hoja_movimiento.cell(row=9, column=7).number_format = '#,##0'
            premio_sin_iva_total = 0
            premio_con_iva_total = 0
            # Iterar sobre los vehículos y agregar la información al archivo Excel existente
            for vehiculo in vehiculos_movimiento:
                
                hoja_movimiento.cell(row=next_row, column=1, value=vehiculo.id).font = font
                hoja_movimiento.cell(row=next_row, column=2, value=vehiculo.marca).font = font
                hoja_movimiento.cell(row=next_row, column=3, value=vehiculo.descripcion).font = font
                hoja_movimiento.cell(row=next_row, column=4, value=vehiculo.anio).font = font
                hoja_movimiento.cell(row=next_row, column=5, value=vehiculo.patente).font = font
                
                hoja_movimiento.cell(row=next_row, column=6, value=vehiculo.premio_sin_iva).font = font
                hoja_movimiento.cell(row=next_row, column=6).number_format = '#,##0'
                hoja_movimiento.cell(row=next_row, column=7, value=vehiculo.premio_con_iva).font = font
                hoja_movimiento.cell(row=next_row, column=7).number_format = '#,##0'
                # Aplicar color de relleno 
                hoja_movimiento.cell(row=next_row, column=6).fill = relleno
                hoja_movimiento.cell(row=next_row, column=7).fill = relleno
                
                # Aplicar bordes
                for col in range(1, 8):
                    hoja_movimiento.cell(row=next_row, column=col).border = bordes
                
                next_row += 1
                num_vehiculo += 1
                premio_sin_iva_total += vehiculo.premio_sin_iva
                premio_con_iva_total += vehiculo.premio_con_iva
                
            hoja_movimiento.cell(row=next_row+1, column=6, value=premio_sin_iva_total).font = font
            hoja_movimiento.cell(row=next_row+1, column=6).number_format = '#,##0'
            hoja_movimiento.cell(row=next_row+1, column=7, value=premio_con_iva_total).font = font
            hoja_movimiento.cell(row=next_row+1, column=7).number_format = '#,##0'
            hoja_movimiento.cell(row=next_row+1, column=6).fill = relleno
            hoja_movimiento.cell(row=next_row+1, column=7).fill = relleno
            hoja_movimiento.cell(row=next_row+1, column=6).border = bordes
            hoja_movimiento.cell(row=next_row+1, column=7).border = bordes
 

        # Crear una respuesta HTTP con el archivo Excel duplicado adjunto
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={flota.cliente.nombre_cliente}.xlsx'
        workbook.save(response)
        workbook.close()

        return response
        

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class FlotasView(View):
    def get(self, request, *args, **kwargs):
        # Obtener el mes seleccionado desde la URL
        selected_month = request.GET.get("month")
        
        # Obtener el primer día del mes seleccionado
        if selected_month:
            selected_month = int(selected_month)
            start_date = datetime(datetime.now().year, selected_month, 1)
            end_date = datetime(datetime.now().year, selected_month + 1, 1) if selected_month < 12 else datetime(datetime.now().year + 1, 1, 1)
            
            # Filtrar las cobranzas para el mes seleccionado
            flotas = Flota.objects.filter(created__gte=start_date, created__lt=end_date).order_by('-created')
        else:
            # Si no se selecciona un mes, muestra todas las cobranzas
            flotas = Flota.objects.all().order_by('-created')
        
        flotas_paginadas = Paginator(flotas, 30)
        page_number = request.GET.get("page")
        filter_pages = flotas_paginadas.get_page(page_number)
        clientes = Cliente.objects.all()
        context = {
            'clientes': clientes,
            'flotas': flotas, 
            'pages': filter_pages,

        }
        return render(request, 'flotas/flotas.html', context)
    def post(self, request, *args, **kwargs):
        # Obtener los datos del formulario directamente desde request.POST
        created = datetime.now()
        numero_flota = request.POST.get('numero')
        poliza = request.POST.get('poliza')
        cliente_id = request.POST.get('cliente')
        
        
        cliente = Cliente.objects.get(pk=cliente_id)
        # Crear una nueva instancia de Flota
        nueva_flota = Flota(
            created = created,
            numero_flota = numero_flota,
            poliza = poliza,
            cliente = cliente,
            
        )
        try:
            # Intenta crear el nuevo elemento
            nueva_flota.save()
            messages.success(request, 'El elemento se creó exitosamente.')
        except Exception as e:
            # Si hay un error al crear el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')
        
        return redirect('flotas')

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class EliminarFlotaView(View):
    def get(self, request, flota_id):
        flota = get_object_or_404(Flota, id=flota_id)

        context = {
            'flota': flota,
        }
        return redirect('flotas')

    def post(self, request, flota_id):
        flota = get_object_or_404(Flota, id=flota_id)
        
        try:
            # Intenta guardar la eliminación del elemento
            flota.delete()
            messages.success(request, 'El elemento se eliminó exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar el elemento. Detalles: {str(e)}')

        return redirect('flotas')

# Flotas
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class DetalleFlotaView(View):
    def get(self, request, flota_id, movimiento_id=None, *args, **kwargs):
        selected_month = request.GET.get("month")
        patente = request.GET.get("patente")
        movimiento_id = request.GET.get("movimiento_id") or movimiento_id  # Obtener movimiento_id de GET si está presente

        # Obtener la flota
        flota = get_object_or_404(Flota, id=flota_id)
        # Obtener todos los movimientos vinculados a esa flota
        movimientos = Movimiento.objects.filter(flota=flota).order_by('fecha_alta_op')

        vehiculos = VehiculoFlota.history.none()  # Inicializar vacío

        if movimiento_id:
            # Si hay un movimiento_id específico, obtener ese movimiento
            movimiento = get_object_or_404(Movimiento, id=movimiento_id)
            print("TOTALES DEL MOVIMIENTO: ")
            print("PRIMA TEC TOTAL: ", movimiento.prima_tec_total)
            print("PRIMA PZA TOTAL: ", movimiento.prima_pza_total)
            print("PREMIO SIN IVA TOTAL: ", movimiento.premio_sin_iva_total)
            print("PREMIO CON IVA TOTAL: ", movimiento.premio_con_iva_total)
            if patente:
                vehiculos = VehiculoFlota.history.filter(movimiento=movimiento, patente__icontains=patente).reverse()
            else:
                vehiculos = VehiculoFlota.history.filter(movimiento=movimiento).reverse()
        else:
            if patente:
                vehiculos = VehiculoFlota.objects.filter(flota=flota, patente__icontains=patente).reverse()
            else:
                vehiculos = VehiculoFlota.objects.filter(flota=flota).reverse()
                
        prima_tecnica_total = 0
        prima_pza_total = 0
        premio_sin_iva_total = 0
        premio_con_iva_total = 0
        
        prima_pza_todos = 0
        premio_sin_iva_todos = 0
        premio_con_iva_todos = 0
        for movimiento in movimientos:
            prima_pza_todos += movimiento.prima_pza_total
            premio_sin_iva_todos += movimiento.premio_sin_iva_total
            premio_con_iva_todos += movimiento.premio_con_iva_total
        print("TOTAL MOVIMIENTOS PRIMA PZA: ", prima_pza_todos)
        print("TOTAL MOVIMIENTOS PREMIO SIN IVA: ", premio_sin_iva_todos)
        print("TOTAL MOVIMIENTOS PREMIO CON IVA: ", premio_con_iva_todos)
        for vehiculo in vehiculos:
            prima_tecnica_total += vehiculo.prima_tecnica
            prima_pza_total += vehiculo.prima_pza
            premio_sin_iva_total += vehiculo.premio_sin_iva
            premio_con_iva_total += vehiculo.premio_con_iva
        print(prima_tecnica_total)
        print(prima_pza_total)
        print(premio_sin_iva_total)
        print(premio_con_iva_total)
        page_number = request.GET.get("page")
        vehiculos_paginados = Paginator(vehiculos, 30)
        filter_pages = vehiculos_paginados.get_page(page_number)

        context = {
            'flota': flota,
            'vehiculos': vehiculos,
            'movimientos': movimientos,
            'pages': filter_pages,
        }
        
        return render(request, 'flotas/detalle_flota.html', context)
    @transaction.atomic
    def post(self, request, flota_id, *args, **kwargs):
        lista_errores = []
        flota = Flota.objects.get(id=flota_id)
        cliente = flota.cliente
        if 'aplicar_diferencia' in request.POST:
            flota.aplicar_diferencia()
        if 'descargar_excel' in request.POST:
            
            # Nombre del archivo que quieres descargar
            file_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel', 'modelo_ejemplo.xlsx')

            # Abre el archivo y lee su contenido
            with open(file_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=modelo_ejemplo.xlsx'
                return response
        if "editar_movimiento" in request.POST:
            movimiento_id = request.POST.get('movimiento_id')
            movimiento = get_object_or_404(Movimiento, id=movimiento_id)
            
            numero_endoso = request.POST.get('numero_endoso')
            motivo_endoso = request.POST.get('motivo_endoso')
            numero_orden = request.POST.get('numero_orden')
            fecha_alta_op = request.POST.get('fecha_alta_op')
            
            movimiento.numero_endoso = numero_endoso
            movimiento.motivo_endoso = motivo_endoso
            movimiento.numero_orden = numero_orden
            movimiento.fecha_alta_op = fecha_alta_op
            
            try:
                # Intenta guardar
                movimiento.save()
                messages.success(request, 'El elemento se actualizó exitosamente.')
            except Exception as e:
                # Si hay un error captura la excepción
                messages.error(request, f'Error: No se pudo actualizar el elemento. Detalles: {str(e)}')
            return redirect('detalle_flota', flota_id = flota_id)
        
        if "comparar_totales" in request.POST:
            # Obtener la instancia de la Flota por su id
            flota = Flota.objects.get(pk=flota_id)

            # Acceder al cliente relacionado
            cliente = flota.cliente
            
            file1 = request.FILES.get('file1')
            workbook = openpyxl.load_workbook(file1)
            try:
                # Intenta comparar y agregar el porcentaje de diferencia al mov
                comparar_totales(workbook, flota_id, cliente)
                messages.success(request, 'Se compararon los totales exitosamente.')
                
            except Exception as e:
                
                messages.error(request, f'Error: No se pudo comparar los totales: {str(e)}')
            return redirect('detalle_flota', flota_id=flota_id)
        
        if "calcular_excel" in request.POST:
            start_time = time.time()
            # Llama a la función para obtener el diccionario de tarifas
            tarifas_dict = get_tarifas()
            
            created = datetime.now()
            api_manager = ApiManager()
            nombre_movimiento = request.POST.get('nombre_movimiento')
            tipo_movimiento = request.POST.get('tipo_movimiento')
            # Mapeo de tipos a cadenas
            tipo_mapping = {
                "1": 'Combinado',
                "2": 'Alta',
                "3": 'Baja',
            }
            # Obtener la instancia de la Flota por su id
            flota = Flota.objects.get(pk=flota_id)

            # Acceder al cliente relacionado
            cliente = flota.cliente
            # Obtener el valor correspondiente o 'No especificado' si el tipo no está en el diccionario
            tipo_string = tipo_mapping.get(tipo_movimiento, 'No especificado')
            
            # Variables para el seguimiento del número de orden actual
            numero_orden_actual = None
            nuevo_movimiento = None
             
            file1 = request.FILES.get('file1')
            workbook = openpyxl.load_workbook(file1)
            
            fuente_datos = request.POST.get('fuente_datos')
            formato_datos = request.POST.get('formato_datos')
            # Si la fuente de datos será info auto, autenticarse
            if fuente_datos == 'info_auto':
                access_token = api_manager.get_valid_access_token()
            
            # Guardar la hoja de cálculo actualizada
            # output = BytesIO()
            # workbook.save(output)
            # output.seek(0)
            # Guardar el último movimiento después de salir del bucle
            if nuevo_movimiento:
                nuevo_movimiento.save()
                
            if cliente.nombre_cliente == 'ROEMMERS SAICF':
                importar_datos_roemmers_saicf(workbook, flota_id, fuente_datos, cliente)
            elif cliente.nombre_cliente == 'ROFINA SAICF':
                importar_datos_rofina_saicf(workbook, flota_id, fuente_datos, cliente)
            elif cliente.nombre_cliente == 'ROEMMERS ALBERTO GUILLERMO':
                importar_datos_roemmers_alberto_guillermo(workbook, flota_id, fuente_datos, cliente)
            elif cliente.nombre_cliente == 'GANADERA SANTA ISABEL':
                importar_datos_ganadera_santa_isabel(workbook, flota_id,fuente_datos, cliente)
            # Crear una respuesta HTTP con el archivo adjunto
            #response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            #response['Content-Disposition'] = f'attachment; filename=resultados_actualizados.xlsx'
            workbook.close()
            end_time = time.time()
            execution_time = end_time - start_time
            print(f"Tiempo de ejecución: {execution_time} segundos")
            messages.success(request, 'El elemento se importó exitosamente.')
            return redirect('detalle_flota', flota_id=flota_id)
        return redirect('detalle_flota', flota_id=flota_id)
    
    #def calcular_datos_con_access_token(self, access_token):
        
# Tarifas flotas
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class DeleteAllTarifasFlotasView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Intenta guardar la eliminación del elemento
            TarifaFlota.objects.all().delete()  # Elimina todos los registros de Tarifas
            messages.success(request, 'Los datos se eliminaron exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar los datos. Detalles: {str(e)}')
        
        return redirect('tarifas_flotas')
    
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class TarifasFlotasView(View):
    
    def get(self, request, *args, **kwargs):

        tarifas = TarifaFlota.objects.all()
    
        tarifas_paginadas = Paginator(tarifas, 30)
        page_number = request.GET.get("page")
        filter_pages = tarifas_paginadas.get_page(page_number)

        context = {
            'tarifas': tarifas, 
            'pages': filter_pages,

        }
        return render(request, 'tarifas_flotas/tarifas_flotas.html', context)
    def obtener_datos_por_rango(self, pagina, inicio, fin):
        datos = []
        for row in range(inicio, fin + 1):
            fila = []
            for col in range(1, pagina.max_column + 1):
                valor = pagina.cell(row=row, column=col).value
                fila.append(valor)
            datos.append((row, fila))  # Se añade el número de fila junto con los datos
        return datos
    
    def obtener_antiguedad_y_tipo_vehiculo(self, titulo):
        mapeo_vehiculos = {
            'Autos hasta 5 años de antigüedad': ('5', 'AUTO'),
            'Autos de 6 a 10 años de antigüedad': ('6 A 10', 'AUTO'),
            'Autos de mas de 10 años de antigüedad': ('MÁS DE 10', 'AUTO'),
            'Pick ups Clase A hasta 5 años de antigüedad': ('5', 'PICK UP CLASE A'),
            'Pick ups Clase A de 6 a 10 años de antigüedad': ('6 A 10', 'PICK UP CLASE A'),
            'Pick ups Clase A de mas de 10 años de antigüedad': ('MÁS DE 10', 'PICK UP CLASE A'),
            'Pick ups 4x4 hasta 5 años de antigüedad': ('5', 'PICK UP 4X4'),
            'Pick ups 4x4 de 6 a 10 años de antigüedad': ('6 A 10', 'PICK UP 4X4'),
            'Pick ups 4x4 de mas de 10 años de antigüedad': ('MÁS DE 10', 'PICK UP 4X4'),
            'Pick ups Clase B hasta 5 años de antigüedad': ('5', 'PICK UP CLASE B'),
            'Pick ups Clase B de 6 a 10 años de antigüedad': ('6 A 10', 'PICK UP CLASE B'),
            'Pick ups Clase B de mas de 10 años de antigüedad': ('MÁS DE 10', 'PICK UP CLASE B')
        }
        return mapeo_vehiculos.get(titulo, (None, None))  # Si el título no está en el diccionario, devuelve None

    def mapear_nombre_zona(self, nombre_zona):
        zonas = {
            'ZONA CAPITAL FEDERAL, ZONA NORTE (Florida, Olivos, San Fdo, San Isidro, Tigre, Vte Lopez)': 'ZONA CAPITAL FEDERAL, ZONA NORTE',
            'ZONA SUR (Villa Dominico, Burzaco, Adrogue, Banfield, Bernal, Burzaco, Lanus, Monte Grande, Quilmes, Luis Guillon, Lomas de Zamora, San Justo)': 'ZONA SUR',
            'ZONA OESTE (Ramos Mejia, Moreno, Castelar, Ituzaingo)': 'ZONA OESTE',
            'ZONA LA PLATA (La Plata, Berisso)': 'ZONA LA PLATA',
            'ZONA CORDOBA (Capital)/ Catamarca/ La Rioja / San Juan / San Luis / Santiago del Estero / Formosa / Jujuy / Salta / Tucumán': 'ZONA CORDOBA',
            'ZONA MENDOZA (Capital) Neuquén / La Pampa / Chubut / Río Negro / Santa Cruz': 'ZONA MENDOZA',
            'ZONA SANTA FE (Capital) / Chaco / Entre Ríos / Corrientes / Misiones': 'ZONA SANTA FE',
            'ZONA ROSARIO': 'ZONA ROSARIO',
            'ZONA MAR DEL PLATA': 'ZONA MAR DEL PLATA'
        }
        return zonas.get(nombre_zona, nombre_zona)  # Si el nombre de la zona no está en el diccionario, devuelve el nombre original

    def guardar_datos_por_condicion(self, datos, nombre_zona):
        for num_fila, fila in datos:
            titulo = fila[0]
        
            antiguedad, tipo_vehiculo = self.obtener_antiguedad_y_tipo_vehiculo(titulo)
            zona = self.mapear_nombre_zona(nombre_zona)
            # Cobertura básica
            tasa = fila[1]
            prima_rc = fila[2]
            tipo_cobertura = 'COB BASICA'
            # Guardar en la BD
            tarifa = TarifaFlota(
                titulo=titulo,
                antiguedad=antiguedad,
                tipo_vehiculo=tipo_vehiculo,
                zona=zona,
                tasa=tasa,
                prima_rc_anual=prima_rc,
                tipo_cobertura=tipo_cobertura,
            )
            tarifa.save()
            # Cobertura clásica
            tasa = fila[3]
            prima_rc = fila[4]
            tipo_cobertura = 'COB CLASICA'
            # Guardar en la BD
            tarifa = TarifaFlota(
                titulo=titulo,
                antiguedad=antiguedad,
                tipo_vehiculo=tipo_vehiculo,
                zona=zona,
                tasa=tasa,
                prima_rc_anual=prima_rc,
                tipo_cobertura=tipo_cobertura,
            )
            tarifa.save()
            # Cobertura póliza 10
            tasa = fila[5]
            prima_rc = fila[6]
            tipo_cobertura = 'COB POLIZA 10'
            # Guardar en la BD
            tarifa = TarifaFlota(
                titulo=titulo,
                antiguedad=antiguedad,
                tipo_vehiculo=tipo_vehiculo,
                zona=zona,
                tasa=tasa,
                prima_rc_anual=prima_rc,
                tipo_cobertura=tipo_cobertura,
            )
            tarifa.save()
            # Cobertura todo riesgo
            tasa = fila[7]
            prima_rc = fila[8]
            tipo_cobertura = 'COB TODO AUTO'
            # Guardar en la BD
            tarifa = TarifaFlota(
                titulo=titulo,
                antiguedad=antiguedad,
                tipo_vehiculo=tipo_vehiculo,
                zona=zona,
                tasa=tasa,
                prima_rc_anual=prima_rc,
                tipo_cobertura=tipo_cobertura,
            )
            tarifa.save()
    def post(self, request, *args, **kwargs):
        
        if "delete_data" in request.POST:
            TarifaFlota.objects.all().delete()
            return redirect('tarifas_flotas')
        if "importar_excel" in request.POST:
            file1 = request.FILES.get('file1')
            wb = openpyxl.load_workbook(file1)
            
            primer_pagina = wb.active
            
            primer_pagina = wb['Table 1']
            segunda_pagina = wb['Table 2']
            tercer_pagina = wb['Table 3']
            cuarta_pagina = wb['Table 4']
            quinta_pagina = wb['Table 5']

            # Obtener información de zonas
            zona1 = primer_pagina['B1':'I1']
            zona2 = primer_pagina['B17':'I17']
            zona3 = segunda_pagina['B1':'I1']
            zona4 = segunda_pagina['B16':'I16']
            zona5 = tercer_pagina['B1':'I1']
            zona6 = tercer_pagina['B16':'I16']
            zona7 = cuarta_pagina['B1':'I1']
            zona8 = cuarta_pagina['B16':'I16']
            zona9 = quinta_pagina['B1':'I1']
            
            # Guardar nombres de zonas
            nombre_zona1 = [celda.value for fila in zona1 for celda in fila]
            nombre_zona2 = [celda.value for fila in zona2 for celda in fila]
            nombre_zona3 = [celda.value for fila in zona3 for celda in fila]
            nombre_zona4 = [celda.value for fila in zona4 for celda in fila]
            nombre_zona5 = [celda.value for fila in zona5 for celda in fila]
            nombre_zona6 = [celda.value for fila in zona6 for celda in fila]
            nombre_zona7 = [celda.value for fila in zona7 for celda in fila]
            nombre_zona8 = [celda.value for fila in zona8 for celda in fila]
            nombre_zona9 = [celda.value for fila in zona9 for celda in fila]
            
            # Guardar datos de zona 1
            datos_zona_1 = self.obtener_datos_por_rango(primer_pagina, 5, 16)
            self.guardar_datos_por_condicion(datos_zona_1, nombre_zona1[0])
            
            # Guardar datos de zona 2
            datos_zona_2 = self.obtener_datos_por_rango(primer_pagina, 20, 31)
            self.guardar_datos_por_condicion(datos_zona_2, nombre_zona2[0])
            
            # Guardar datos de zona 3
            datos_zona_3 = self.obtener_datos_por_rango(segunda_pagina, 4, 15)
            self.guardar_datos_por_condicion(datos_zona_3, nombre_zona3[0])
            
            # Guardar datos de zona 4
            datos_zona_4 = self.obtener_datos_por_rango(segunda_pagina, 19, 30)
            self.guardar_datos_por_condicion(datos_zona_4, nombre_zona4[0])

            # Guardar datos de zona 5
            datos_zona_5 = self.obtener_datos_por_rango(tercer_pagina, 4, 15)
            self.guardar_datos_por_condicion(datos_zona_5, nombre_zona5[0])
            
            # Guardar datos de zona 6
            datos_zona_6 = self.obtener_datos_por_rango(tercer_pagina, 19, 30)
            self.guardar_datos_por_condicion(datos_zona_6, nombre_zona6[0])
            
            # Guardar datos de zona 7
            datos_zona_7 = self.obtener_datos_por_rango(cuarta_pagina, 4, 15)
            self.guardar_datos_por_condicion(datos_zona_7, nombre_zona7[0])
            
            # Guardar datos de zona 8
            datos_zona_8 = self.obtener_datos_por_rango(cuarta_pagina, 19, 30)
            self.guardar_datos_por_condicion(datos_zona_8, nombre_zona8[0])
            
            # Guardar datos de zona 9
            datos_zona_9 = self.obtener_datos_por_rango(quinta_pagina, 4, 15)
            self.guardar_datos_por_condicion(datos_zona_9, nombre_zona9[0])

                
        # Obtén los datos del formulario directamente desde request.POST
        titulo = request.POST.get('titulo')
        zona = request.POST.get('zona')
        tipo_vehiculo = request.POST.get('tipo_vehiculo')
        antiguedad = request.POST.get('antiguedad')
        tipo_cobertura = request.POST.get('tipo_cobertura')
        tasa = request.POST.get('tasa')
        prima_rc_anual = request.POST.get('prima_rc_anual')

        


        if not prima_rc_anual or not tasa:
         
            return redirect('tarifas_flotas')  # Redirige al usuario nuevamente al formulario

        # Crea una nueva instancia de Flota y guárdala en la base de datos
        nueva_tarifa_flota = TarifaFlota(
            titulo = titulo,
            zona = zona,
            tipo_vehiculo = tipo_vehiculo,
            antiguedad = antiguedad,
            tipo_cobertura = tipo_cobertura,
            tasa = tasa,
            prima_rc_anual = prima_rc_anual
            # Agrega los otros campos del formulario aquí
        )
        nueva_tarifa_flota.save()

        # Redirige a la página de flotas o realiza alguna otra acción que desees
        return redirect('tarifas_flotas')


@method_decorator(login_required, name='dispatch')   
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class DetalleTarifaFlotaView(View):
    def get(self, request, tarifa_id):
        tarifa = get_object_or_404(TarifaFlota, id=tarifa_id)
        # Formatear los decimales para poder mostrarlos
        tasa_formatted = "{:.3f}".format(tarifa.tasa).replace(',', '.')
        prima_formatted = "{:.3f}".format(tarifa.prima_rc_anual).replace(',', '.')
        context = {
            'tarifa': tarifa,
            'tasa_formatted': tasa_formatted,
            'prima_formatted': prima_formatted,
        }
        return render(request, 'tarifas_flotas/detalle_tarifa_flota.html', context)

    def post(self, request, tarifa_id):
        tarifa = get_object_or_404(TarifaFlota, id=tarifa_id)
        # Obtén los datos del formulario directamente desde request.POST
        titulo = request.POST.get('titulo')
        zona = request.POST.get('zona')
        tipo_vehiculo = request.POST.get('tipo_vehiculo')
        antiguedad = request.POST.get('antiguedad')
        tipo_cobertura = request.POST.get('tipo_cobertura')
        tasa = request.POST.get('tasa')
        prima_rc_anual = request.POST.get('prima_rc_anual')
        
        # Actualiza los campos de la tarifa con los datos del formulario
        tarifa.titulo = titulo
        tarifa.zona = zona
        tarifa.tipo_vehiculo = tipo_vehiculo
        tarifa.antiguedad = antiguedad
        tarifa.tipo_cobertura = tipo_cobertura
        tarifa.tasa = tasa
        tarifa.prima_rc_anual = prima_rc_anual
        try:
            # Intenta guardar la actualización del elemento
            tarifa.save()
            messages.success(request, 'El elemento se actualizó exitosamente.')
        except Exception as e:
            # Si hay un error al actualizar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo actualizar el elemento. Detalles: {str(e)}')
        return redirect('tarifas_flotas')

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class EliminarTarifaFlotaView(View):
    def get(self, request, tarifa_id):
        tarifa = get_object_or_404(TarifaFlota, id=tarifa_id)
        # Muestra una página de confirmación para la eliminación
        context = {
            'tarifa': tarifa,
        }
        return redirect('tarifas_flotas')

    def post(self, request, tarifa_id):
        tarifa = get_object_or_404(TarifaFlota, id=tarifa_id)

        try:
            # Intenta guardar la eliminación del elemento
            tarifa.delete()
            messages.success(request, 'El elemento se eliminó exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar el elemento. Detalles: {str(e)}')


        return redirect('tarifas_flotas')


# Cobranzas
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoAvanzado').exists() or user.is_staff), name='dispatch')
class CobranzasView(View):
    def get(self, request, *args, **kwargs):
         # Obtén el mes seleccionado desde la URL
        selected_month = request.GET.get("month")
        
        # Obtiene el primer día del mes seleccionado
        if selected_month:
            selected_month = int(selected_month)
            start_date = datetime(datetime.now().year, selected_month, 1)
            end_date = datetime(datetime.now().year, selected_month + 1, 1) if selected_month < 12 else datetime(datetime.now().year + 1, 1, 1)
            
            # Filtra las cobranzas para el mes seleccionado
            vencimientos = Vencimiento.objects.filter(fecha_vencimiento__gte=start_date, fecha_vencimiento__lt=end_date).order_by('-fecha_vencimiento')
        else:
            # Si no se selecciona un mes, muestra todas las cobranzas
            vencimientos = Vencimiento.objects.all().order_by('-fecha_vencimiento')
        
        cobranzas_paginadas = Paginator(vencimientos, 30)
        page_number = request.GET.get("page")
        filter_pages = cobranzas_paginadas.get_page(page_number)

        context = {
            'vencimientos': vencimientos,  # Pasar las cobranzas al contexto
            'pages': filter_pages,

        }
        return render(request, 'cobranzas/cobranzas.html', context)
    
    
    def post(self, request, *args, **kwargs):
        if "delete_data" in request.POST:
            Vencimiento.objects.all().delete()  # Elimina todos los registros de Cobranza
            return redirect('cobranzas')
        
        
        selected_month = int(request.POST.get('month'))
        selected_year = int(request.POST.get('year'))
        file1 = request.FILES.get('file1')
        workbook = openpyxl.load_workbook(file1)
        sheet = workbook.active

        data = []  # Lista para almacenar las filas
        for row in sheet.iter_rows(min_row=4, values_only=True):
            asegurado, riesgo, productor, cliente, poliza, endoso, cuota, fecha_vencimiento, moneda, importe, saldo, forma_pago, factura = row
            # Si hay fecha cambia el formato al necesario por Django
            if fecha_vencimiento is not None:
                fecha_vencimiento = fecha_vencimiento.replace("/", "-")
                fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%d-%m-%Y")

                # Verifica si la fecha de vencimiento está en el mes y año seleccionados por el usuario
                if fecha_vencimiento.month == selected_month and fecha_vencimiento.year == selected_year:
                    data.append(row)  # Agrega la fila a la lista

                    # Procesa las filas de 100 en 100
                    if len(data) == 5:
                        self.process_data(data, selected_month, selected_year)
                        data = []  # Reinicia la lista para el siguiente lote

        # Procesa cualquier lote restante (menos de 100 filas)
        if data:
            self.process_data(data, selected_month, selected_year)
        
        context = {
            
        } 

        return redirect('cobranzas')
    
    def process_data(self, data, selected_month, selected_year):
        for row in data:
            asegurado, riesgo, productor, cliente, poliza, endoso, cuota, fecha_vencimiento, moneda, importe, saldo, forma_pago, factura = row
            # Si hay fecha de vencimiento cambia el formato al necesario por Django
            if fecha_vencimiento is not None:
                fecha_vencimiento = fecha_vencimiento.replace("/", "-")
                
                fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%d-%m-%Y")
                
                if fecha_vencimiento.month == selected_month and fecha_vencimiento.year == selected_year:
                    Vencimiento.objects.create(
                        asegurado=asegurado,
                        riesgo=riesgo,
                        productor=productor,
                        cliente=cliente,
                        poliza=poliza,
                        endoso=endoso,
                        cuota=cuota,
                        fecha_vencimiento=fecha_vencimiento,
                        moneda=moneda,
                        importe=importe,
                        saldo=saldo,
                        forma_pago=forma_pago,
                        factura=factura
                    )

# Vencimientos
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoAvanzado').exists() or user.is_staff), name='dispatch')
class VencimientosView(View):
    def get(self, request, *args, **kwargs):
         # Obtén el mes seleccionado desde la URL
        selected_month = request.GET.get("month")
        
        # Obtiene el primer día del mes seleccionado
        if selected_month:
            selected_month = int(selected_month)
            start_date = datetime(datetime.now().year, selected_month, 1)
            end_date = datetime(datetime.now().year, selected_month + 1, 1) if selected_month < 12 else datetime(datetime.now().year + 1, 1, 1)
            
            # Filtra las cobranzas para el mes seleccionado
            vencimientos = Vencimiento.objects.filter(fecha_vencimiento__gte=start_date, fecha_vencimiento__lt=end_date).order_by('-fecha_vencimiento')
        else:
            # Si no se selecciona un mes, muestra todas las cobranzas
            vencimientos = Vencimiento.objects.all().order_by('-fecha_vencimiento')
        
        cobranzas_paginadas = Paginator(vencimientos, 30)
        page_number = request.GET.get("page")
        filter_pages = cobranzas_paginadas.get_page(page_number)

        context = {
            'vencimientos': vencimientos,  # Pasar las cobranzas al contexto
            'pages': filter_pages,

        }
        return render(request, 'vencimientos/vencimientos.html', context)
    
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        if "delete_data" in request.POST:
            Vencimiento.objects.all().delete()  # Elimina todos los registros de Cobranza
            return redirect('vencimientos')
        
        
        selected_month = int(request.POST.get('month'))
        selected_year = int(request.POST.get('year'))
        file1 = request.FILES.get('file1')
        workbook = openpyxl.load_workbook(file1)
        sheet = workbook.active

        data = []  # Lista para almacenar las filas
        for row in sheet.iter_rows(min_row=4, values_only=True):
            asegurado, riesgo, productor, cliente, poliza, endoso, cuota, fecha_vencimiento, moneda, importe, saldo, forma_pago, factura = row
            # Si hay fecha cambia el formato al necesario por Django
            if fecha_vencimiento is not None:
                fecha_vencimiento = fecha_vencimiento.replace("/", "-")
                fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%d-%m-%Y")

                # Verifica si la fecha de vencimiento está en el mes y año seleccionados por el usuario
                if fecha_vencimiento.month == selected_month and fecha_vencimiento.year == selected_year:
                    data.append(row)  # Agrega la fila a la lista

                    # Procesa las filas de 100 en 100
                    if len(data) == 5:
                        self.process_data(data, selected_month, selected_year)
                        data = []  # Reinicia la lista para el siguiente lote

        # Procesa cualquier lote restante (menos de 100 filas)
        if data:
            self.process_data(data, selected_month, selected_year)
        
        context = {
            
        } 

        return redirect('vencimientos')
    
    def process_data(self, data, selected_month, selected_year):
        for row in data:
            asegurado, riesgo, productor, cliente, poliza, endoso, cuota, fecha_vencimiento, moneda, importe, saldo, forma_pago, factura = row
            # Si hay fecha de vencimiento cambia el formato al necesario por Django
            if fecha_vencimiento is not None:
                fecha_vencimiento = fecha_vencimiento.replace("/", "-")
                
                fecha_vencimiento = datetime.strptime(fecha_vencimiento, "%d-%m-%Y")
                
                if fecha_vencimiento.month == selected_month and fecha_vencimiento.year == selected_year:
                    Vencimiento.objects.create(
                        asegurado=asegurado,
                        riesgo=riesgo,
                        productor=productor,
                        cliente=cliente,
                        poliza=poliza,
                        endoso=endoso,
                        cuota=cuota,
                        fecha_vencimiento=fecha_vencimiento,
                        moneda=moneda,
                        importe=importe,
                        saldo=saldo,
                        forma_pago=forma_pago,
                        factura=factura
                    )
                    
# Localidades
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoAvanzado').exists() or user.is_staff), name='dispatch')
class LocalidadesView(View):
    def get(self, request, *args, **kwargs):
        localidades = Localidad.objects.all()
        cobranzas_paginadas = Paginator(localidades, 30)
        page_number = request.GET.get("page")
        filter_pages = cobranzas_paginadas.get_page(page_number)

        context = {
            'localidades': localidades, 
            'pages': filter_pages,

        }
        return render(request, 'localidades/localidades.html', context)
    def post(self, request, *args, **kwargs):
        lista_errores = []
        context = {
            'errores': lista_errores
        }
        if 'delete_data' in request.POST:
            Localidad.objects.all().delete()
            return redirect('localidades')
        if 'normalizar_nombres' in request.POST:
            start_time = time.time()
            localidades = Localidad.objects.all()
            for localidad in localidades:
                # Cambiar nombres a mayúsculas y sin tilde
                nombre_localidad_normalizado = unidecode(localidad.nombre_localidad.upper())
                nombre_municipio_normalizado = unidecode(localidad.nombre_municipio.upper())
                nombre_provincia_normalizado = unidecode(localidad.nombre_provincia.upper())
                # Actualiza los nombres de la localidad, municipio y provincia en la base de datos
                localidad.nombre_localidad = nombre_localidad_normalizado
                localidad.nombre_municipio = nombre_municipio_normalizado
                localidad.nombre_provincia = nombre_provincia_normalizado
                localidad.save()
            end_time = time.time()
            execution_time = end_time - start_time
            print(f"Tiempo de ejecución, {execution_time} segundos")
        if 'importar_excel' in request.POST:
            start_time = time.time()
            file1 = request.FILES.get('file1')

            # Envuelve el archivo en un TextIOWrapper para manejar la codificación
            csv_file_wrapper = TextIOWrapper(file1.file, encoding='utf-8')

            # Usa DictReader para obtener un diccionario por fila
            csv_reader = csv.DictReader(csv_file_wrapper)

            # Itera sobre las filas del CSV
            for row_number, row in enumerate(csv_reader, start=2):
                municipio = row.get('municipio_nombre', '')
                localidad = row.get('nombre', '')
                provincia = row.get('provincia_nombre', '')
                zona = row.get('zona_nombre', '')

                # Crea el objeto Localidad
                Localidad.objects.create(
                    nombre_localidad=localidad,
                    nombre_municipio=municipio,
                    nombre_provincia=provincia,
                    zona=zona
                )
            end_time = time.time()
            execution_time = end_time - start_time
            print(f"Tiempo de ejecución, {execution_time} segundos")
        return redirect('localidades')

# Seguros de creditos

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class CreditosView(View):
    def get(self, request, *args, **kwargs):
        
        asegurados = AseguradoCredito.objects.all()
        
        asegurados_paginados = Paginator(asegurados, 30)
        page_number = request.GET.get("page")
        filter_pages = asegurados_paginados.get_page(page_number)
        context = {

            'asegurados': asegurados, 
            'pages': filter_pages,

        }
        return render(request, 'creditos/creditos.html', context)
    def post(self, request, *args, **kwargs):
        # Obtén los datos del formulario directamente desde request.POST
        
        nombre = request.POST.get('nombre')
        cuit = request.POST.get('cuit')
        direccion = request.POST.get('direccion')
        provincia = request.POST.get('provincia')
        num_poliza = request.POST.get('num_poliza')
        producto = request.POST.get('producto')
        vigencia_desde = request.POST.get('vigencia_desde')
        vigencia_hasta = request.POST.get('vigencia_hasta')
    
        nuevo_asegurado = AseguradoCredito(
            
            nombre_asegurado=nombre,
            cuit=cuit,
            direccion=direccion,
            provincia=provincia,
            numero_poliza=num_poliza,
            producto=producto,
            fecha_vigencia_desde=vigencia_desde,
            fecha_vigencia_hasta=vigencia_hasta,

        )
        try:
            # Intenta crear el nuevo elemento
            nuevo_asegurado.save()
            messages.success(request, 'El elemento se creó exitosamente.')
        except Exception as e:
            # Si hay un error al crear el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo crear el elemento. Detalles: {str(e)}')

        # Redirige, incluyendo los mensajes en el contexto
        return HttpResponseRedirect(request.path_info)

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class DetalleCreditoView(View):
    def get(self, request, asegurado_id, *args, **kwargs):
       
        # Obtener el asegurado
        asegurado = get_object_or_404(AseguradoCredito, id=asegurado_id)
        
        nominados = CoberturaNominada.objects.filter(asegurado=asegurado)
        innominados = CoberturaInnominada.objects.filter(asegurado=asegurado)

        page_number_nominados = request.GET.get("page")
        nominados_paginados = Paginator(nominados, 30)
        filter_pages_nominados = nominados_paginados.get_page(page_number_nominados)

        page_number_innominados = request.GET.get("page")
        innominados_paginados = Paginator(innominados, 30)
        filter_pages_innominados = innominados_paginados.get_page(page_number_innominados)

        context = {
            'asegurado': asegurado,
            'nominados': nominados,
            'pages_nominados': filter_pages_nominados,
            'pages_innominados': filter_pages_innominados,
        }
        
        return render(request, 'creditos/detalle_credito.html', context)
    @transaction.atomic
    def post(self, request, asegurado_id, *args, **kwargs):
        lista_errores = []
        asegurado = AseguradoCredito.objects.get(id=asegurado_id)

        if 'generar_reporte' in request.POST:
            año = int(request.POST.get('año'))
            mes = int(request.POST.get('mes'))
            
            fecha_formateada = f"{mes:02d}/{año}"
            fecha_completa_formateada = f"01/{mes:02d}/{año}"
            print(fecha_completa_formateada)
            # Obtener los datos filtrados
            datos_solicitudes_cobertura = obtener_datos_solicitudes_cobertura(fecha_completa_formateada, asegurado)
            
            datos_clientes_nuevos = obtener_datos_clientes_sin_cobertura(fecha_completa_formateada, asegurado)
            
            
            # Renderizar el template con los datos
            html_string = render_to_string('creditos/reporte_template.html', {
            'asegurado': asegurado,
            'fecha_formateada': fecha_formateada,
            'datos_solicitudes_cobertura': datos_solicitudes_cobertura,
            'datos_clientes_nuevos': datos_clientes_nuevos,
            
            })
            
            # Generar el PDF
            html = HTML(string=html_string)
            pdf = html.write_pdf()

            # Devolver el PDF como respuesta HTTP
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="reporte.pdf"'
            return response
            
        if "importar_nominados" in request.POST:
            # Obtener la instancia del asegurado
            asegurado = AseguradoCredito.objects.get(pk=asegurado_id)

            
            file1 = request.FILES.get('file1')
            df = pd.read_excel(file1)
            try:
                # Intenta importar datos de nominados
                
                cargar_datos_nominados(df, asegurado)
                messages.success(request, 'Se importaron los datos exitosamente.')
                
            except Exception as e:
                
                messages.error(request, f'Error: No se pudo importar los datos: {str(e)}')
            return redirect('detalle_credito', asegurado_id=asegurado_id)
        
        if "importar_innominados" in request.POST:
            # Obtener la instancia del asegurado
            asegurado = AseguradoCredito.objects.get(pk=asegurado_id)

            
            file2 = request.FILES.get('file1')
            
            df = pd.read_excel(file2)
            try:
                # Intenta importar datos de nominados
                
                cargar_datos_innominados(df, asegurado)
                messages.success(request, 'Se importaron los datos exitosamente.')
                
            except Exception as e:
                
                messages.error(request, f'Error: No se pudo importar los datos: {str(e)}')
            return redirect('detalle_credito', asegurado_id=asegurado_id)
        
        return redirect('detalle_credito', asegurado_id=asegurado_id)


@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class EliminarAseguradoCreditoView(View):
    def get(self, request, asegurado_id):
        asegurado = get_object_or_404(AseguradoCredito, id=asegurado_id)

        context = {
            'asegurado': asegurado
        }
        return redirect('creditos')

    def post(self, request, asegurado_id):
        asegurado = get_object_or_404(AseguradoCredito, id=asegurado_id)
        
        try:
            # Intenta guardar la eliminación del elemento
            asegurado.delete()
            messages.success(request, 'El elemento se eliminó exitosamente.')
        except Exception as e:
            # Si hay un error al eliminar el elemento, captura la excepción
            messages.error(request, f'Error: No se pudo eliminar el elemento. Detalles: {str(e)}')

        return redirect('creditos')        
        

# Buscar vehículo 

@login_required
def autocomplete_marcas(request):
    term = request.GET.get('term', '')
    marcas = MarcaInfoAuto.objects.filter(nombre__icontains=term).values('id', 'nombre')
    return JsonResponse(list(marcas), safe=False)

@login_required
def obtener_vehiculos_por_marca(request, marca_id):
    # Lógica para obtener vehículos por marca (ajusta esto según tus modelos)
    vehiculos = VehiculoInfoAuto.objects.filter(marca__id=marca_id).values('id', 'descripcion')

    # Devuelve la lista de vehículos en formato JSON
    return JsonResponse(list(vehiculos), safe=False)

@login_required
def buscar_vehiculo_por_codigo(request, codigo):
    if request.method == 'GET':
      
        vehiculo = VehiculoInfoAuto.objects.get(codigo=codigo)
        precios = PrecioAnual.objects.filter(vehiculo=vehiculo)
        # Obtener precios del vehículo
        precios = PrecioAnual.objects.filter(vehiculo=vehiculo).order_by('-anio')

        # Crear una lista de años y precios para enviar en la respuesta JSON
        anios_precios = [{'anio': precio.anio, 'precio': precio.precio} for precio in precios]

        data = {
                'codigo': vehiculo.codigo,
                'marca': vehiculo.marca.nombre,
                'descripcion': vehiculo.descripcion,
                'tipo': vehiculo.tipo_vehiculo,
                'nacionalidad': vehiculo.nacionalidad,
                'precios': anios_precios,
                
                # ... otros campos que quieras incluir ...
            }
        if vehiculo.precio_okm:
            data['okm'] = vehiculo.precio_okm

        return JsonResponse(data)

@login_required
def obtener_datos_vehiculo(request, vehiculo_id):
        
        try:
            vehiculo = VehiculoInfoAuto.objects.get(pk=vehiculo_id)
            precios = PrecioAnual.objects.filter(vehiculo=vehiculo)
            # Obtener precios del vehículo
            precios = PrecioAnual.objects.filter(vehiculo=vehiculo).order_by('-anio')

            # Crear una lista de años y precios para enviar en la respuesta JSON
            anios_precios = [{'anio': precio.anio, 'precio': precio.precio} for precio in precios]
            
            data = {
                'codigo': vehiculo.codigo,
                'marca': vehiculo.marca.nombre,
                'descripcion': vehiculo.descripcion,
                'tipo': vehiculo.tipo_vehiculo,
                'nacionalidad': vehiculo.nacionalidad,
                'precios': anios_precios,
                
                # ... otros campos que quieras incluir ...
            }
            if vehiculo.precio_okm:
                data['okm'] = vehiculo.precio_okm
            
            return JsonResponse(data)
        except VehiculoInfoAuto.DoesNotExist:
            return JsonResponse({'error': 'Vehículo no encontrado'}, status=404)

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoBasico').exists() or user.is_staff), name='dispatch')
class BuscarVehiculoView(View):
    def get(self, request, *args, **kwargs):
        marcas = MarcaInfoAuto.objects.order_by('nombre')
        
        context = {
            'marcas': marcas,
        }
        return render(request, 'info_auto/buscar_vehiculo.html', context)
# Vehículos info auto
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(lambda user: user.groups.filter(name='PermisoAvanzado').exists() or user.is_staff), name='dispatch')
class VehiculosInfoAutoView(View):
    def get(self, request, *args, **kwargs):
        vehiculos = VehiculoInfoAuto.objects.order_by('marca__nombre')
        vehiculos_paginados = Paginator(vehiculos, 30)
        page_number = request.GET.get("page")
        filter_pages = vehiculos_paginados.get_page(page_number)

        context = {
            'vehiculos': vehiculos,
            'pages': filter_pages,

        }
        return render(request, 'info_auto/vehiculos_info_auto.html', context)

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        if 'delete_data' in request.POST:
            VehiculoInfoAuto.objects.all().delete()
            MarcaInfoAuto.objects.all().delete()
            return redirect('vehiculos')
        
        if 'importar_excel' in request.POST:
            start_time = time.time()
            archivo = request.FILES.get('file1')
            formato = request.POST.get('formato')

            if formato == 'csv':
                datos = TextIOWrapper(archivo, encoding='utf-8')
                datos_csv = csv.reader(datos)
            elif formato == 'xlsx':
                datos = BytesIO(archivo.read())
                libro_excel = openpyxl.load_workbook(datos)
                hoja_activa = libro_excel.active
                datos_excel = list(hoja_activa.iter_rows(values_only=True))
            else:
                return HttpResponse("Formato de archivo no válido")

            lista_vehiculos_nuevos = []
            lista_precios_anuales = []
            vehiculos_actualizar = []

            
            marcas = {marca.nombre: marca for marca in MarcaInfoAuto.objects.all()}
            vehiculos_existentes = {vehiculo.codigo: vehiculo for vehiculo in VehiculoInfoAuto.objects.all()}
            precios_anuales_existentes = {vehiculo.id: {precio.anio: precio for precio in PrecioAnual.objects.filter(vehiculo_id=vehiculo.id)} for vehiculo in vehiculos_existentes.values()}

            for index, fila in enumerate(datos_csv, start=1) if formato == 'csv' else enumerate(datos_excel, start=1):
                if index < 4:
                    continue

                cod, marca_nombre, descripcion, nacionalidad, tipo, okm, *precios = fila

                marca = marcas.get(marca_nombre)
                if not marca:
                    marca, _ = MarcaInfoAuto.objects.get_or_create(nombre=marca_nombre)
                    marcas[marca_nombre] = marca

                if isinstance(okm, str) and okm.isdigit():
                    precio_okm = Decimal(okm.replace(',', '.'))
                else:
                    precio_okm = Decimal(0)

                vehiculo, created = VehiculoInfoAuto.objects.get_or_create(
                    codigo=cod,
                    defaults={
                        'marca': marca,
                        'descripcion': descripcion,
                        'nacionalidad': nacionalidad,
                        'tipo_vehiculo': tipo,
                        'precio_okm': precio_okm,
                    }
                )

                if not created:
                    if vehiculo.descripcion != descripcion or vehiculo.nacionalidad != nacionalidad or vehiculo.tipo_vehiculo != tipo or vehiculo.precio_okm != precio_okm:
                        vehiculo.descripcion = descripcion
                        vehiculo.nacionalidad = nacionalidad
                        vehiculo.tipo_vehiculo = tipo
                        vehiculo.precio_okm = precio_okm
                        vehiculos_actualizar.append(vehiculo)

                for i, year in enumerate(range(2024, 2003, -1)):
                    precio_str = precios[i] if i < len(precios) else None
                    if precio_str and precio_str != '':
                        precio_decimal = Decimal(str(precio_str).replace(',', '.'))
                        if vehiculo.id in precios_anuales_existentes and year in precios_anuales_existentes[vehiculo.id]:
                            precio_anual_existente = precios_anuales_existentes[vehiculo.id][year]
                            if precio_anual_existente.precio != precio_decimal:
                                precio_anual_existente.precio = precio_decimal
                                precio_anual_existente.save()
                        else:
                            lista_precios_anuales.append(PrecioAnual(vehiculo=vehiculo, anio=year, precio=precio_decimal))

            VehiculoInfoAuto.objects.bulk_update(vehiculos_actualizar, ['descripcion', 'nacionalidad', 'tipo_vehiculo', 'precio_okm'])
            PrecioAnual.objects.bulk_create(lista_precios_anuales)
            end_time = time.time()
            execution_time = end_time - start_time
            print(execution_time)
            return redirect('vehiculos')
        
# Autenticación
class SignOutView(View):
    def get(self, request, *args, **kwargs):
        logout(request)
        return redirect('login')

class SignInView(View):
    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('inicio')
        return render(request, 'login.html', {
            'form': AuthenticationForm()
        })

    def post(self, request, *args, **kwargs):
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('inicio')
        
        return render(request, 'login.html', {
            'form': form,
            'error': 'El nombre de usuario o la contraseña no existen',
        })
        
# 1687, max lines